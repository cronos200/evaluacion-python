import openpyxl


estudiantes = {}

for i in range(3):
    nombre = input("Ingresa el nombre del estudiante: ")
    nota = float(input(f"Ingresa la nota de {nombre}: "))
    estudiantes[nombre] = nota

try:  
    libro = openpyxl.load_workbook('mi_primer_excel.xlsx')
    hoja = libro.active
except FileNotFoundError:
   
    print("No se encontró 'mi_primer_excel.xlsx'. Se creará uno nuevo.")
    libro = openpyxl.Workbook()
    hoja = libro.active

if hoja['A1'].value is None:
    hoja['A1'] = 'Estudiante'
    hoja['B1'] = 'Clasificación'

fila = hoja.max_row + 1  
for estudiante, nota in estudiantes.items():
    hoja[f'A{fila}'] = estudiante
    if nota > 70:
        hoja[f'B{fila}'] = 'Bueno'
    else:
        hoja[f'B{fila}'] = 'Regular'
    fila += 1

libro.save('mi_primer_excel.xlsx')

print("¡Datos guardados en 'mi_primer_excel.xlsx'!")
